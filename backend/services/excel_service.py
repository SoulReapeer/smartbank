import io
from datetime import datetime
import pandas as pd

def generate_excel(transactions, account_number=None, filename_prefix='transactions'):
    rows = []
    for tx in transactions:
        if tx.transaction_type == 'deposit':
            direction = 'IN'
        elif tx.transaction_type == 'withdrawal':
            direction = 'OUT'
        elif tx.transaction_type == 'transfer':
            direction = 'OUT' if tx.sender_account == account_number else 'IN'
        else:
            direction = '—'
        rows.append({
            'Date': tx.timestamp.strftime('%Y-%m-%d'),
            'Time': tx.timestamp.strftime('%H:%M:%S'),
            'Type': tx.transaction_type.capitalize(),
            'Direction': direction,
            'Amount (BDT)': round(tx.amount, 2),
            'Category': tx.category or 'Others',
            'Reference': tx.reference or '',
            'Sender Account': tx.sender_account or '',
            'Receiver Account': tx.receiver_account or '',
        })

    df = pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=['Date','Time','Type','Direction','Amount (BDT)','Category','Reference','Sender Account','Receiver Account'])

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Transactions', index=False, startrow=3)
        ws = writer.sheets['Transactions']
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        navy_fill  = PatternFill('solid', fgColor='1E3A8A')
        light_fill = PatternFill('solid', fgColor='F8FAFC')
        green_fill = PatternFill('solid', fgColor='D1FAE5')
        red_fill   = PatternFill('solid', fgColor='FEE2E2')
        thin = Border(
            left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0'))

        ws['A1'] = '💳 SmartBank — Transaction Export'
        ws['A1'].font = Font(bold=True, size=14, color='1E3A8A')
        ws['A2'] = f'Generated: {datetime.utcnow().strftime("%B %d, %Y %H:%M")} UTC' + (f'  |  Account: {account_number}' if account_number else '')
        ws['A2'].font = Font(size=9, color='64748B')
        ws.merge_cells('A1:I1')
        ws.merge_cells('A2:I2')

        header_row = 4
        for ci, cn in enumerate(df.columns, 1):
            cell = ws.cell(row=header_row, column=ci)
            cell.fill = navy_fill
            cell.font = Font(bold=True, color='FFFFFF', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin

        for ri, row_data in enumerate(rows, header_row + 1):
            for ci in range(1, len(df.columns) + 1):
                cell = ws.cell(row=ri, column=ci)
                cell.border = thin
                cell.alignment = Alignment(vertical='center')
                if ri % 2 == 0:
                    cell.fill = light_fill
                if ci == 4:
                    if cell.value == 'IN':
                        cell.fill = green_fill; cell.font = Font(bold=True, color='065F46')
                    elif cell.value == 'OUT':
                        cell.fill = red_fill; cell.font = Font(bold=True, color='991B1B')
                if ci == 5:
                    cell.number_format = '#,##0.00'

        for i, w in enumerate([12,10,14,12,16,14,36,18,18], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[4].height = 22
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buffer.seek(0)
    fname = f"{filename_prefix}_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return buffer.read(), fname
